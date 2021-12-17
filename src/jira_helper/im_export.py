"""
Import and export of JIRA results
"""
from openpyxl import \
        Workbook, \
        load_workbook
from openpyxl.worksheet.dimensions import \
        ColumnDimension, \
        DimensionHolder
from openpyxl.styles import \
        Alignment, \
        Border, \
        Side, \
        Font, \
        PatternFill
from openpyxl.utils import \
        get_column_letter


# Header font
FONT_HEADER = Font(
        bold=True
)
# Header background
FILL_HEADER = PatternFill(
        fill_type='solid',
        fgColor='00CCFFFF'
)

# Text alignment to 'word wrap'
ALIGNMENT = Alignment(
        wrap_text=True
)

# Border syle
BORDER_THIN = Side(
        style='thin',
        color='00000000'
)
BORDER = Border(
        left=BORDER_THIN,
        right=BORDER_THIN,
        top=BORDER_THIN,
        bottom=BORDER_THIN
)


def export_query(results, fields, server, filename):
    """
    Export a query result as .xslx
    """
    # create workboot/worksheet
    work_book = Workbook()
    work_sheet = work_book.active
    work_sheet.title = 'JIRA query result'


    # Populate data structure
    header = [ 'key' ] + fields
    work_sheet.append(header)
    for result in results:
        row = []
        for column in header:
            row.append(result[column])
        work_sheet.append(row)

    # Turn key into a link to the issue
    for row in work_sheet.iter_rows(min_row=2):
        row[0].hyperlink=f'{server}/browse/{row[0].value}'
        row[0].style='Hyperlink'

    # Apply styles
    for row in work_sheet:
        for cell in row:
            cell.alignment = ALIGNMENT
            cell.border = BORDER

    # Apply styles to header
    for column in work_sheet[1]:
        column.font = FONT_HEADER
        column.fill = FILL_HEADER

    # Freeze top row (header)
    work_sheet.freeze_panes = 'A2'

    # Add autofilter for header
    work_sheet.auto_filter.ref=work_sheet.dimensions

    # auto with for columns
    dimension_holder = DimensionHolder(worksheet=work_sheet)
    for column in range(work_sheet.min_column, work_sheet.max_column + 1):
        dimension_holder[get_column_letter(column)] = ColumnDimension(
                work_sheet,
                min=column,
                max=column,
                width=20
        )
    work_sheet.column_dimensions = dimension_holder

    # Write file to disk
    work_book.save(filename)


def import_query(filename):
    """
    Import .xslx
    """
    work_book = load_workbook(filename = filename)
    work_sheet = work_book.active

    issues = []

    for row in work_sheet.iter_rows(min_row=2):
        entries = {}
        for i, entry in enumerate(row):
            entries[work_sheet[1][i].value] = entry.value
        issues.append(entries)

    return issues
