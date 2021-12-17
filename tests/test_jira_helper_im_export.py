"""
Unit tests for jira_helper/im_export.py
"""
import unittest
import settings
from openpyxl import load_workbook
from jira_helper import im_export


class TestJiraHelperImExportMethods(unittest.TestCase):
    """
    Test class for jira_helper/im_export.py
    """

    def test_export_query(self):
        """
        Test export_query(results, fields, filename)
        """
        results = [
            {
                'key': 'PLN-1',
                'assignee': 'testuser',
                'status': 'New',
                'summary': 'Test story'
            },
            {
                'key': 'PLN-2',
                'assignee': 'other user',
                'status': 'Done',
                'summary': 'Another story'
            }
        ]
        fields = [
                'summary',
                'status',
                'assignee'
        ]
        filename = 'test_result.xlsx'

        # Export data to xslx
        im_export.export_query(
                results,
                fields,
                settings.SERVER,
                filename
        )

        # Read back exported data
        work_book = load_workbook(filename)
        work_sheet = work_book.active

        # Check header, remember that rows/columns start to count from 1
        # First entry in header is 'key'
        self.assertEqual(work_sheet.cell(row=1, column=1).value, 'key')
        # Now check if other fields are in header
        for i, field in enumerate(fields, 1):
            self.assertEqual(
                    work_sheet.cell(row=1, column=i+1).value,
                    field
            )

        # Check rows
        for i, row in enumerate(results, 1):
            # First entry is 'key'
            self.assertEqual(
                    work_sheet.cell(row=i+1, column=1).value,
                    row['key']
            )
            # Get order from fields
            for j, col in enumerate(fields, 1):
                self.assertEqual(
                        work_sheet.cell(row=i+1, column=j+1).value,
                        row[col]
                )

    def test_import_query(self):
        """
        Test import_query(filename)
        """
        filename = 'test_result.xlsx'

        result = im_export.import_query(filename)

        self.assertIsNotNone(result)
